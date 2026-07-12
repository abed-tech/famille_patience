"""Génération de rapports PDF, Excel et CSV."""
import csv
import io
from datetime import datetime, timedelta

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from apps.members.models import (
    Member, MemberStatus, FamilyPole, ChurchPole, ChurchDepartment, Profession,
)
from apps.events.models import Event
from apps.attendance.models import Attendance
from apps.accounts.models import User, UserRole

PERIOD_LABELS = {
    "daily": "Journalier",
    "weekly": "Hebdomadaire",
    "monthly": "Mensuel",
    "yearly": "Annuel",
    "custom": "Personnalisé",
}

MODULE_LABELS = {
    "all": "Complet",
    "members": "Membres",
    "registrations": "Inscriptions",
    "attendance": "Présences",
    "events": "Événements",
    "referrers": "Référents",
    "counsellors": "Conseillers",
    "poles": "Pôles FP",
    "departments": "Départements",
    "professions": "Professions",
}

HEADER_FILL = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def resolve_period(period="monthly", start_date=None, end_date=None):
    today = timezone.now().date()
    if period == "custom":
        start = _parse_date(start_date)
        end = _parse_date(end_date)
        if start and end and start <= end:
            return start, end
        return today - timedelta(days=30), today
    if period == "daily":
        return today, today
    if period == "weekly":
        return today - timedelta(days=6), today
    if period == "yearly":
        return today.replace(month=1, day=1), today
    return today - timedelta(days=29), today


def _style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _members_qs(start, end, module):
    qs = Member.objects.all().order_by("-registration_date")
    if module == "registrations":
        return qs.filter(registration_date__gte=start, registration_date__lte=end)
    if module in ("members", "all"):
        return qs.filter(registration_date__lte=end)
    return qs.none()


def _attendance_qs(start, end):
    return Attendance.objects.filter(
        scanned_at__date__gte=start,
        scanned_at__date__lte=end,
    ).select_related("member", "event")


def _events_qs(start, end):
    return Event.objects.filter(date__gte=start, date__lte=end).order_by("-date")


def _referrers_qs():
    return User.objects.filter(role=UserRole.REFERRER, is_active=True).order_by("last_name")


def _counsellors_qs():
    return User.objects.filter(role=UserRole.COUNSELLOR, is_active=True).order_by("last_name")


def get_report_preview(period="monthly", module="all", start_date=None, end_date=None):
    start, end = resolve_period(period, start_date, end_date)
    members = Member.objects.filter(registration_date__lte=end)
    new_members = members.filter(registration_date__gte=start, registration_date__lte=end)
    attendances = _attendance_qs(start, end)
    events = _events_qs(start, end)

    summary = [
        {"label": "Membres actifs", "value": Member.objects.filter(status=MemberStatus.ACTIVE).count()},
        {"label": "Nouvelles inscriptions", "value": new_members.count()},
        {"label": "Présences enregistrées", "value": attendances.filter(is_present=True).count()},
        {"label": "Événements", "value": events.count()},
        {"label": "Référents actifs", "value": _referrers_qs().count()},
        {"label": "Conseillers actifs", "value": _counsellors_qs().count()},
        {"label": "Pôles FP", "value": FamilyPole.objects.filter(is_active=True).count()},
        {"label": "Départements", "value": ChurchDepartment.objects.filter(is_active=True).count()},
    ]

    rows = []
    if module in ("all", "members", "registrations"):
        for m in new_members[:15] if module == "registrations" else members.filter(
            registration_date__gte=start
        )[:15]:
            rows.append({
                "type": "member",
                "label": m.full_name,
                "detail": m.member_number,
                "meta": str(m.registration_date),
            })
    elif module == "attendance":
        for a in attendances.filter(is_present=True)[:15]:
            rows.append({
                "type": "attendance",
                "label": a.member.full_name,
                "detail": a.event.name,
                "meta": a.scanned_at.strftime("%d/%m/%Y %H:%M"),
            })
    elif module == "events":
        for e in events[:15]:
            rows.append({
                "type": "event",
                "label": e.name,
                "detail": e.location or "—",
                "meta": str(e.date),
            })
    elif module == "referrers":
        for r in _referrers_qs()[:15]:
            count = Member.objects.filter(referrer=r).count()
            rows.append({"type": "referrer", "label": r.full_name, "detail": r.email, "meta": f"{count} membres"})
    elif module == "counsellors":
        for c in _counsellors_qs()[:15]:
            count = Member.objects.filter(counsellor=c).count()
            rows.append({"type": "counsellor", "label": c.full_name, "detail": c.email, "meta": f"{count} membres"})

    return {
        "period": period,
        "period_label": PERIOD_LABELS.get(period, period),
        "module": module,
        "module_label": MODULE_LABELS.get(module, module),
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "start_display": start.strftime("%d/%m/%Y"),
        "end_display": end.strftime("%d/%m/%Y"),
        "summary": summary,
        "preview_rows": rows,
        "generated_at": timezone.now().isoformat(),
    }


def _new_sheet(wb, title):
    if not wb.sheetnames:
        ws = wb.create_sheet(title)
    else:
        ws = wb.create_sheet(title)
    return ws


def _append_members_sheet(wb, start, end, module):
    ws = _new_sheet(wb, "Membres")
    ws.append(["N°", "Nom", "Prénom", "Téléphone", "Statut", "Inscription"])
    _style_header_row(ws)
    qs = _members_qs(start, end, "registrations" if module == "registrations" else "members")
    for m in qs:
        ws.append([m.member_number, m.last_name, m.first_name, m.phone_primary, m.status, str(m.registration_date)])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18


def _append_attendance_sheet(wb, start, end):
    ws = wb.create_sheet("Présences")
    ws.append(["Membre", "Événement", "Date scan", "Présent"])
    _style_header_row(ws)
    for a in _attendance_qs(start, end):
        ws.append([
            a.member.full_name, a.event.name,
            a.scanned_at.strftime("%d/%m/%Y %H:%M"),
            "Oui" if a.is_present else "Non",
        ])


def _append_events_sheet(wb, start, end):
    ws = wb.create_sheet("Événements")
    ws.append(["Nom", "Date", "Lieu", "Statut", "Présences"])
    _style_header_row(ws)
    for e in _events_qs(start, end):
        ws.append([e.name, str(e.date), e.location, e.status, e.attendances.filter(is_present=True).count()])


def _append_staff_sheet(wb, title, qs, role_label):
    ws = wb.create_sheet(title)
    ws.append(["Nom", "Email", "Téléphone", "Membres assignés"])
    _style_header_row(ws)
    for u in qs:
        field = "referrer" if role_label == "referrer" else "counsellor"
        count = Member.objects.filter(**{field: u}).count()
        ws.append([u.full_name, u.email, u.phone or "", count])


def _append_poles_sheet(wb):
    ws = wb.create_sheet("Pôles FP")
    ws.append(["Nom", "Description", "Actif", "Membres intéressés"])
    _style_header_row(ws)
    for p in FamilyPole.objects.all().order_by("name"):
        count = Member.objects.filter(interested_family_pole=p).count()
        ws.append([p.name, p.description or "", "Oui" if p.is_active else "Non", count])


def _append_departments_sheet(wb):
    ws = wb.create_sheet("Départements")
    ws.append(["Nom", "Pôle église", "Description", "Actif"])
    _style_header_row(ws)
    for d in ChurchDepartment.objects.select_related("pole").order_by("pole__name", "name"):
        ws.append([d.name, d.pole.name, d.description or "", "Oui" if d.is_active else "Non"])


def _append_professions_sheet(wb):
    ws = wb.create_sheet("Professions")
    ws.append(["Profession", "Actif", "Membres"])
    _style_header_row(ws)
    for p in Profession.objects.all().order_by("name"):
        ws.append([p.name, "Oui" if p.is_active else "Non", Member.objects.filter(profession_ref=p).count()])


def _append_summary_sheet(wb, start, end):
    members = Member.objects.filter(registration_date__gte=start, registration_date__lte=end)
    attendances = _attendance_qs(start, end)
    events = _events_qs(start, end)
    ws = wb.create_sheet("Résumé")
    ws.append(["Indicateur", "Valeur"])
    _style_header_row(ws)
    rows = [
        ("Total membres actifs", Member.objects.filter(status=MemberStatus.ACTIVE).count()),
        ("Nouvelles inscriptions", members.count()),
        ("Événements", events.count()),
        ("Présences", attendances.filter(is_present=True).count()),
        ("Référents", _referrers_qs().count()),
        ("Conseillers", _counsellors_qs().count()),
        ("Pôles FP actifs", FamilyPole.objects.filter(is_active=True).count()),
        ("Départements actifs", ChurchDepartment.objects.filter(is_active=True).count()),
    ]
    for label, value in rows:
        ws.append([label, value])


def generate_excel_report(period="monthly", module="all", start_date=None, end_date=None):
    start, end = resolve_period(period, start_date, end_date)
    wb = Workbook()
    wb.remove(wb.active)

    sheets = {
        "all": ["summary", "members", "attendance", "events", "referrers", "counsellors", "poles", "departments"],
        "members": ["members", "summary"],
        "registrations": ["members", "summary"],
        "attendance": ["attendance", "summary"],
        "events": ["events", "summary"],
        "referrers": ["referrers", "summary"],
        "counsellors": ["counsellors", "summary"],
        "poles": ["poles", "summary"],
        "departments": ["departments", "summary"],
        "professions": ["professions", "summary"],
    }
    for key in sheets.get(module, sheets["all"]):
        if key == "summary":
            _append_summary_sheet(wb, start, end)
        elif key == "members":
            _append_members_sheet(wb, start, end, module)
        elif key == "attendance":
            _append_attendance_sheet(wb, start, end)
        elif key == "events":
            _append_events_sheet(wb, start, end)
        elif key == "referrers":
            _append_staff_sheet(wb, "Référents", _referrers_qs(), "referrer")
        elif key == "counsellors":
            _append_staff_sheet(wb, "Conseillers", _counsellors_qs(), "counsellor")
        elif key == "poles":
            _append_poles_sheet(wb)
        elif key == "departments":
            _append_departments_sheet(wb)
        elif key == "professions":
            _append_professions_sheet(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_csv_report(period="monthly", module="all", start_date=None, end_date=None):
    start, end = resolve_period(period, start_date, end_date)
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow([f"Rapport {MODULE_LABELS.get(module, module)} — Famille Patience"])
    writer.writerow([f"Période : {start.strftime('%d/%m/%Y')} — {end.strftime('%d/%m/%Y')}"])
    writer.writerow([])

    preview = get_report_preview(period, module, start_date, end_date)
    writer.writerow(["Indicateur", "Valeur"])
    for item in preview["summary"]:
        writer.writerow([item["label"], item["value"]])
    writer.writerow([])

    if module in ("all", "members", "registrations"):
        writer.writerow(["N°", "Nom", "Prénom", "Téléphone", "Statut", "Inscription"])
        qs = _members_qs(start, end, "registrations" if module == "registrations" else "members")
        for m in qs:
            writer.writerow([m.member_number, m.last_name, m.first_name, m.phone_primary, m.status, m.registration_date])
    elif module == "attendance":
        writer.writerow(["Membre", "Événement", "Date scan", "Présent"])
        for a in _attendance_qs(start, end):
            writer.writerow([a.member.full_name, a.event.name, a.scanned_at.strftime("%d/%m/%Y %H:%M"), "Oui" if a.is_present else "Non"])
    elif module == "events":
        writer.writerow(["Nom", "Date", "Lieu", "Statut", "Présences"])
        for e in _events_qs(start, end):
            writer.writerow([e.name, e.date, e.location, e.status, e.attendances.filter(is_present=True).count()])
    elif module == "referrers":
        writer.writerow(["Nom", "Email", "Téléphone", "Membres"])
        for u in _referrers_qs():
            writer.writerow([u.full_name, u.email, u.phone or "", Member.objects.filter(referrer=u).count()])
    elif module == "counsellors":
        writer.writerow(["Nom", "Email", "Téléphone", "Membres"])
        for u in _counsellors_qs():
            writer.writerow([u.full_name, u.email, u.phone or "", Member.objects.filter(counsellor=u).count()])
    elif module == "poles":
        writer.writerow(["Nom", "Description", "Actif", "Membres intéressés"])
        for p in FamilyPole.objects.all().order_by("name"):
            writer.writerow([p.name, p.description or "", "Oui" if p.is_active else "Non", Member.objects.filter(interested_family_pole=p).count()])
    elif module == "departments":
        writer.writerow(["Nom", "Pôle église", "Description", "Actif"])
        for d in ChurchDepartment.objects.select_related("pole").order_by("pole__name", "name"):
            writer.writerow([d.name, d.pole.name, d.description or "", "Oui" if d.is_active else "Non"])
    elif module == "professions":
        writer.writerow(["Profession", "Actif", "Membres"])
        for p in Profession.objects.all().order_by("name"):
            writer.writerow([p.name, "Oui" if p.is_active else "Non", Member.objects.filter(profession_ref=p).count()])

    out = io.BytesIO()
    out.write(buffer.getvalue().encode("utf-8-sig"))
    out.seek(0)
    return out


def generate_pdf_report(period="monthly", module="all", start_date=None, end_date=None):
    start, end = resolve_period(period, start_date, end_date)
    preview = get_report_preview(period, module, start_date, end_date)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(
        f"<b>Rapport {preview['module_label']} — {preview['period_label']}</b>",
        styles["Title"],
    ))
    elements.append(Paragraph(
        f"Période : {preview['start_display']} — {preview['end_display']}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 16))

    summary_data = [["Indicateur", "Valeur"]] + [[s["label"], str(s["value"])] for s in preview["summary"]]
    table = Table(summary_data, colWidths=[220, 100])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ec4899")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fdf2f8")]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 16))

    if preview["preview_rows"]:
        elements.append(Paragraph("<b>Aperçu des données</b>", styles["Heading2"]))
        detail_data = [["Libellé", "Détail", "Info"]] + [
            [r["label"], r["detail"], r["meta"]] for r in preview["preview_rows"][:25]
        ]
        detail_table = Table(detail_data, colWidths=[180, 180, 120])
        detail_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ec4899")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fdf2f8")]),
        ]))
        elements.append(detail_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer
