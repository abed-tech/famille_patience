"""Rapports de présence par événement (PDF / Excel / JSON)."""
import io
from xml.sax.saxutils import escape

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from apps.members.models import Member, MemberStatus
from apps.attendance.models import Attendance
from apps.attendance.services import scan_mode_label


def _safe_photo_url(request, member):
    if not request or not member.photo:
        return None
    try:
        url = member.photo.url
        if url.startswith(("http://", "https://")):
            return url
        return request.build_absolute_uri(url)
    except Exception:
        return None


def get_event_report_data(event, request=None, counsellor_user=None):
    """Construit les données du rapport pour un événement fermé ou en cours."""
    if counsellor_user:
        members_qs = Member.objects.filter(
            counsellor=counsellor_user, status=MemberStatus.ACTIVE
        ).select_related("referrer")
        member_ids = list(members_qs.values_list("id", flat=True))
        total_expected = len(member_ids)
    else:
        members_qs = Member.objects.filter(status=MemberStatus.ACTIVE).select_related("referrer")
        member_ids = list(members_qs.values_list("id", flat=True))
        total_expected = len(member_ids)

    attendances = (
        Attendance.objects.filter(event=event, member_id__in=member_ids)
        .select_related("member", "member__referrer", "scanned_by")
        .order_by("member__last_name", "member__first_name")
    )

    present_qs = attendances.filter(is_present=True)
    absent_qs = attendances.filter(is_present=False)

    def serialize_row(att):
        member = att.member
        ref = member.referrer
        return {
            "member_id": str(member.id),
            "full_name": member.full_name,
            "photo": _safe_photo_url(request, member),
            "referrer_name": ref.full_name if ref else "—",
            "referrer_id": str(ref.id) if ref else None,
            "scanned_at": att.scanned_at.isoformat() if att.scanned_at else None,
            "scan_mode": scan_mode_label(att) if att.is_present else None,
        }

    present = [serialize_row(a) for a in present_qs]
    absent = [serialize_row(a) for a in absent_qs]

    present_count = len(present)
    absent_count = len(absent)
    rate = round((present_count / total_expected) * 100, 1) if total_expected else 0

    return {
        "event": {
            "id": str(event.id),
            "name": event.name,
            "description": event.description,
            "date": event.date.isoformat(),
            "time": str(event.time)[:5] if event.time else None,
            "location": event.location or "",
            "status": event.status,
            "closed_at": event.updated_at.isoformat() if event.status == "closed" else None,
        },
        "summary": {
            "total_expected": total_expected,
            "present_count": present_count,
            "absent_count": absent_count,
            "attendance_rate": rate,
        },
        "present": present,
        "absent": absent,
        "generated_at": timezone.now().isoformat(),
    }


def generate_event_excel_report(event, request=None):
    data = get_event_report_data(event, request)
    wb = Workbook()
    header_fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")

    ws = wb.active
    ws.title = "Resume"
    ev = data["event"]
    s = data["summary"]
    ws.append(["Rapport d'événement — Famille Patience"])
    ws.append(["Événement", ev["name"]])
    ws.append(["Date", ev["date"]])
    ws.append(["Lieu", ev["location"] or "—"])
    ws.append(["Participants attendus", s["total_expected"]])
    ws.append(["Présents", s["present_count"]])
    ws.append(["Absents", s["absent_count"]])
    ws.append(["Taux de présence", f"{s['attendance_rate']}%"])

    ws2 = wb.create_sheet("Presents")
    ws2.append(["Nom complet", "Référent", "Heure pointage", "Mode"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in data["present"]:
        scanned = row["scanned_at"][:16].replace("T", " ") if row["scanned_at"] else "—"
        ws2.append([row["full_name"], row["referrer_name"], scanned, row["scan_mode"] or "—"])

    ws3 = wb.create_sheet("Absents")
    ws3.append(["Nom complet", "Référent"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in data["absent"]:
        ws3.append([row["full_name"], row["referrer_name"]])

    for sheet in (ws2, ws3):
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_event_pdf_report(event, request=None):
    data = get_event_report_data(event, request)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    ev = data["event"]
    s = data["summary"]
    elements.append(Paragraph(f"<b>Rapport — {escape(ev['name'] or '')}</b>", styles["Title"]))
    elements.append(Paragraph(
        escape(
            f"Date : {ev['date']} · Lieu : {ev['location'] or '—'} · "
            f"Présents : {s['present_count']} · Absents : {s['absent_count']} · "
            f"Taux : {s['attendance_rate']}%"
        ),
        styles["Normal"],
    ))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("<b>Personnes presentes</b>", styles["Heading2"]))
    present_data = [["Nom", "Referent", "Heure", "Mode"]]
    for row in data["present"][:80]:
        scanned = row["scanned_at"][:16].replace("T", " ") if row["scanned_at"] else "—"
        present_data.append([
            row["full_name"] or "—",
            row["referrer_name"] or "—",
            scanned,
            row["scan_mode"] or "—",
        ])
    if len(present_data) == 1:
        present_data.append(["—", "—", "—", "—"])

    pt = Table(present_data, colWidths=[180, 140, 100, 80])
    pt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ec4899")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fdf2f8")]),
    ]))
    elements.append(pt)
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("<b>Personnes absentes</b>", styles["Heading2"]))
    absent_data = [["Nom", "Referent"]]
    for row in data["absent"][:80]:
        absent_data.append([row["full_name"] or "—", row["referrer_name"] or "—"])
    if len(absent_data) == 1:
        absent_data.append(["—", "—"])

    at = Table(absent_data, colWidths=[220, 180])
    at.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ec4899")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fef2f2")]),
    ]))
    elements.append(at)

    doc.build(elements)
    buffer.seek(0)
    return buffer
